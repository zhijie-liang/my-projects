# 工作表排序


from openpyxl import load_workbook


def reorder_excel_sheets(file_path, sheet_order):
    # 加载工作簿
    wb = load_workbook(file_path)

    # 验证所有工作表名是否存在于工作簿中
    missing_sheets = [
        sheet for sheet in sheet_order if sheet not in wb.sheetnames]
    if missing_sheets:
        print("以下工作表在工作簿中不存在，将被忽略：", missing_sheets)

    # 移除不在sheet_order列表中的工作表
    for sheet in wb.sheetnames:
        if sheet not in sheet_order:
            wb.remove(wb[sheet])

    # 按照新的顺序重新添加工作表
    for sheet_name in sheet_order:
        if sheet_name in wb.sheetnames:
            wb._sheets.append(wb._sheets.pop(wb.sheetnames.index(sheet_name)))

    # 保存工作簿
    wb.save(file_path)


# 工作表的新顺序
new_sheet_order = [
    "目录", "tbl_ori_flow_ipv4", "tbl_ori_flow_ipv6", "tbl_perf_tcpflow", "tbl_perf_udp", "tbl_perf_httppageview", "tbl_perf_httprrpair", "tbl_perf_mysqlbiz", "tbl_perf_mysqlrrpair", "tbl_perf_oracle", "tbl_perf_oraclesql", "tbl_perf_dameng", "tbl_perf_damengsql", "tbl_perf_dns", "tbl_perf_pop3conn", "tbl_perf_pop3crpair", "tbl_perf_smtpconn", "tbl_perf_smtpcrpair", "tbl_perf_imapconn", "tbl_perf_imaprrpair", "tbl_perf_ftp", "tbl_perf_ftprrpair", "tbl_perf_ftpdata", "tbl_perf_smb", "tbl_perf_telnetconn", "tbl_perf_telnetrrpair", "tbl_perf_icmp", "tbl_perf_icmprrpair", "tbl_perf_kingbasebiz", "tbl_perf_kingbaserrpair", "tbl_perf_gbasebiz", "tbl_perf_gbaserrpair", "tbl_perf_db2", "tbl_perf_db2sql", "tbl_perf_mssqlbiz", "tbl_perf_mssqlrrpair", "tbl_perf_highgobiz", "tbl_perf_highgorrpair", "tbl_perf_shentongbiz", "tbl_perf_shentongrrpair", "tbl_perf_httpspageview", "tbl_perf_httpsrrpair"
]

# 调用函数
reorder_excel_sheets(
    r"C:\Users\梁智杰\Desktop\核对校验\processed_原始.xlsx", new_sheet_order)
