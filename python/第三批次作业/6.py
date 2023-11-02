# 添加第四列

import openpyxl


def add_merged_column_fixed_v2(input_filepath, output_filepath):
    # Load the workbook and get the active sheet
    workbook = openpyxl.load_workbook(input_filepath)
    sheet = workbook.active

    # Insert a new column at index 4 (which is the position for the fourth column)
    sheet.insert_cols(4)

    # Iterate through the rows in the sheet
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row), start=1):
        # Get the values from the second and third columns
        second_column_value = sheet.cell(row=i, column=2).value
        third_column_value = sheet.cell(row=i, column=3).value

        # Create the merged value
        merged_value = f"{second_column_value}、{third_column_value}"

        # Set the value for the newly inserted fourth column
        sheet.cell(row=i, column=4).value = merged_value

    # Save the workbook to the output filepath
    workbook.save(output_filepath)
    print(f"Processed {input_filepath} and saved to {output_filepath}")


# You can call the function with your filepaths
add_merged_column_fixed_v2(r"C:\Users\梁智杰\Desktop\核对校验\核对.xlsx",
                           r"C:\Users\梁智杰\Desktop\核对校验\核对\核对.xlsx")
