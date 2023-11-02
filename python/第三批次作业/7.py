# 保留三列

import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


def process_workbook(input_path, output_path):
    # 载入Excel工作簿
    workbook = load_workbook(input_path)

    # 遍历所有工作表
    for sheet_name in workbook.sheetnames:
        # 跳过名为“目录”的工作表
        if sheet_name == "目录":
            continue

        sheet = workbook[sheet_name]

        # 删除第三列之后的所有列
        for col in sheet.iter_cols(min_col=4, max_col=sheet.max_column):
            for cell in col:
                sheet.delete_cols(cell.column)

        # 删除所有图片
        for obj in sheet._images:
            sheet._images.remove(obj)

    # 保存处理后的工作簿到输出路径
    workbook.save(output_path)


# 输入和输出路径
input_directory = r"C:\Users\梁智杰\Desktop\核对校验"  # 将此路径替换为您的输入文件所在的目录
output_directory = r"C:\Users\梁智杰\Desktop\核对校验\7"  # 将此路径替换为您希望保存输出文件的目录
input_filename = '原始.xlsx'  # 将此处替换为您的输入文件名
output_filename = '7.xlsx'  # 将此处替换为您的输出文件名

input_path = os.path.join(input_directory, input_filename)
output_path = os.path.join(output_directory, output_filename)

# 处理工作簿
process_workbook(input_path, output_path)
